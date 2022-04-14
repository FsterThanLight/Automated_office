import xlwings as xw
import pypyodbc
import datetime

from PyQt5.QtWidgets import QApplication, QMainWindow, QMessageBox, QFileDialog
from PyQt5.QtCore import QDate
from functions import *
from return_specific_terms_daily import *
import openpyxl
import sys
from auto_fill import Ui_MainWindow

cursor = ''
sht_2 = ''
date = ''
wb_1 = ''
excel_app = ''
from datetime import datetime, timedelta
from openpyxl.styles import Alignment


class Sy(QMainWindow, Ui_MainWindow):
    def __init__(self):
        super().__init__()
        self.setupUi(self)
        # 设置系统日期为今日
        self.date.setDate(QDate.currentDate())
        self.button2.setEnabled(False)
        self.progressBar.setValue(0)
        self.button1.setEnabled(False)


def open_file(sy):
    global wb_1, excel_app
    fileName = QFileDialog.getOpenFileName(sy, "选取文件", os.getcwd(),
                                           "All Files(*);;Text Files(*.txt)")
    print(fileName[0])
    # sy.state.setText('日报名称：'+fileName[0])
    sy.state.setText('日报表加载中...')
    excel_app = xw.App(visible=True, add_book=False)
    excel_app.display_alerts = False
    excel_app.screen_updating = True
    wb_1 = excel_app.books.open(fileName[0])
    sy.state.setText('日报表加载中...')
    sy.state.setText('日报表加载完毕')
    sy.button1.setEnabled(True)
    new_date = datetime.strftime(datetime.strptime(wb_1.sheets[-1].name, '%Y.%m.%d') + timedelta(1), '%Y.%m.%d').split(
        '.')
    sy.date.setDate(QDate(int(new_date[0]), int(new_date[1]), int(new_date[2])))


def accdb(path):
    '''建立与数据库的连接，返回游标'''
    # path=os.path.abspath('..')
    # 取得当前文件目录
    print(path)
    mdb = 'Driver={Microsoft Access Driver (*.mdb,*.accdb)};DBQ=' + path
    # foods_data.mdb是数据库文件
    # 连接字符串
    conn = pypyodbc.win_connect_mdb(mdb)
    # 建立连接
    cursor = conn.cursor()
    return cursor


def extract_data(s_date, e_date, unit, length, project, cursor):
    '''提取有关数据'''
    if project == '梁片预制数量':
        cursor.execute("SELECT count(梁片编码) FROM 所有已浇筑梁片信息 \
        where 浇筑日期>=#" + s_date + "# and 浇筑日期<=#" + e_date + "# and \
        浇筑单位='" + unit + "' and 梁片长度='" + length + "';")
    elif project == '梁片预制产值':
        cursor.execute("SELECT round(sum(总产值)/10000,2) FROM 所有已浇筑梁片信息 \
        where 浇筑日期>=#" + s_date + "# and 浇筑日期<=#" + e_date + "# and \
        浇筑单位='" + unit + "' and 梁片长度='" + length + "';")
    elif project == '梁片安装数量':
        cursor.execute("SELECT count(梁片编码) FROM 所有已安装梁片信息 \
        where 安装日期>=#" + s_date + "# and 安装日期<=#" + e_date + "# and \
        安装单位='" + unit + "' and 梁片长度='" + length + "';")
    elif project == '梁片安装产值':
        cursor.execute("SELECT sum(安装产值) FROM 所有已安装梁片信息 \
        where 安装日期>=#" + s_date + "# and 安装日期<=#" + e_date + "# and \
        安装单位='" + unit + "' and 梁片长度='" + length + "';")
    elif project == '湿接缝长度':
        cursor.execute("SELECT sum(该跨梁长（米）) FROM 已完成湿接缝查询 \
        where 完成日期>=#" + s_date + "# and 完成日期<=#" + e_date + "# and \
        完成单位='" + unit + "';")
    elif project == '湿接缝产值':
        cursor.execute("SELECT round(sum(产值)/10000,2) FROM 已完成湿接缝查询 \
        where 完成日期>=#" + s_date + "# and 完成日期<=#" + e_date + "# and \
        完成单位='" + unit + "';")
    elif project == '湿接缝跨数':
        cursor.execute("SELECT count(跨) FROM 已完成湿接缝汇总查询 \
        where 完成日期>=#" + s_date + "# and 完成日期<=#" + e_date + "# and \
        完成单位='" + unit + "';")
    elif project == '防撞护栏长度':
        cursor.execute("SELECT sum(长度合计) FROM 已完成防撞护栏汇总查询 \
        where 完成日期>=#" + s_date + "# and 完成日期<=#" + e_date + "# and \
        完成单位='" + unit + "';")
    elif project == '防撞护栏产值':
        cursor.execute("SELECT round(sum(价格（元）)/10000,2) FROM 已完成防撞护栏查询 \
        where 完成日期>=#" + s_date + "# and 完成日期<=#" + e_date + "# and \
        完成单位='" + unit + "';")
    elif project == '防撞护栏联数':
        cursor.execute("SELECT count(长度合计) FROM 已完成防撞护栏汇总查询 \
        where 完成日期>=#" + s_date + "# and 完成日期<=#" + e_date + "# and \
        完成单位='" + unit + "'and 长度合计>20 ;")
    elif project == '桥面铺装面积':
        cursor.execute("SELECT sum(平均面积) FROM 已完成桥面铺装查询 \
        where 完成日期>=#" + s_date + "# and 完成日期<=#" + e_date + "# and \
        完成单位='" + unit + "';")
    elif project == '桥面铺装产值':
        cursor.execute("SELECT round(sum(价格（元）)/10000,2) FROM 已完成桥面铺装查询 \
        where 完成日期>=#" + s_date + "# and 完成日期<=#" + e_date + "# and \
        完成单位='" + unit + "';")
    elif project == '桥面铺装联数':
        cursor.execute("SELECT count(联) FROM 已完成桥面铺装查询 \
        where 完成日期>=#" + s_date + "# and 完成日期<=#" + e_date + "# and \
        完成单位='" + unit + "';")
    amount = cursor.fetchall()[0][0]
    if amount is None:
        amount = 0
    return amount


def error_output(cursor, sht_2):
    """计算产值误差"""
    # 计算数据库中的总产值
    cursor.excute('select round(sum(总产值)/10000,2) from 所有已浇筑梁片信息')
    beam_value = cursor.fetchall()[0][0]
    cursor.excute('select sum(安装产值) from 所有已浇筑梁片信息')
    ins_value = cursor.fetchall()[0][0]
    cursor.excute('select round(sum(产值)/10000,2) FROM 已完成湿接缝查询')
    wet_value = cursor.fetchall()[0][0]
    cursor.excute('SELECT round(sum(价格（元）)/10000,2) FROM 已完成防撞护栏查询')
    guardrail_value = cursor.fetchall()[0][0]
    cursor.excute('SELECT round(sum(价格（元）)/10000,2) FROM 已完成桥面铺装查询')
    shop_value = cursor.fetchall()[0][0]
    total_value = beam_value + ins_value + wet_value + guardrail_value + shop_value

    # 获取表中的400章总产值
    total_table_value = sht_2.range('d2').value
    # 计算差值
    error_value = total_table_value - total_value
    return error_value


def main(sy):
    global cursor, date, sht_2, wb_1
    date = 'xx'
    while True:
        '''输入要生成的报表日期，并复制前一天表格'''
        date = sy.date.date().toString('yyyy.M.d')
        print(date)
        # date=input('报表日期(跳过默认今日日期)：')
        if date == '':
            date = datetime.datetime.now().strftime('%#Y.%#m.%#d')
        before_date = str_before_date(date)
        print(before_date)
        sht_1 = wb_1.sheets(before_date)
        try:
            sht_2 = sht_1.copy(name=date, after=sht_1)
            path = sht_2.range('i40').value.replace('"', '')
            break
        except ValueError:
            sy.state.setText(date + '日期信息已经存在，请重新输入。')
            # x=input(date+'日期信息已经存在，请重新输入。')
    # 连接数据库
    try:
        cursor = accdb(path)
        sy.state.setText('成功连接数据库。')
        # print('成功连接数据库。')
    except TypeError:
        sy.state.setText('请在日报表格中‘数据来源’处更新数据库访问路径。')
        # print('未连接数据库，请在日报表格中‘数据来源’处更新数据库访问路径。')
        wb_1.close()
        app.quit()
        sys.exit()
    new_date = date.replace('.', '/')

    # 填写100章产值
    # beam_factory_1_100=input('\n1#梁厂100章今日产值(万元)：')
    try:
        beam_factory_1_100 = float(sy.factory1100.text())
    except ValueError:
        beam_factory_1_100 = 0
    if beam_factory_1_100 == '':
        beam_factory_1_100 = 0
    # beam_factory_2_100=input('\n2#梁厂100章今日产值(万元)：')
    try:
        beam_factory_2_100 = float(sy.factory2100.text())
    except ValueError:
        beam_factory_2_100 = 0
    if beam_factory_2_100 == '':
        beam_factory_2_100 = 0
    sht_2.range('c3').value = beam_factory_1_100
    sht_2.range('c4').value = beam_factory_2_100

    # 填写备注信息
    # from return_specific_terms_daily import return_specific_terms,return_specific_terms_qmx
    j = return_specific_terms(cursor, new_date, '1#智慧梁厂', '浇筑')
    l = return_specific_terms(cursor, new_date, '1#智慧梁厂', '安装')
    s_1 = return_specific_terms_qmx(cursor, new_date, '1#智慧梁厂', '湿接缝名称')
    f_1 = return_specific_terms_qmx(cursor, new_date, '1#智慧梁厂', '防撞护栏名称')
    q_1 = return_specific_terms_qmx(cursor, new_date, '1#智慧梁厂', '桥面铺装名称')
    hui_1 = j + l + s_1 + f_1 + q_1
    hui_1_1 = hui_1.replace("  ", "、").replace(" ", "")
    if hui_1_1 == '':
        hui_1_1 = '无'
    # note_information_1=input('\n1#梁厂今日完成详情：'+hui_1_1)
    sy.factory1400.setText(hui_1_1)
    hui_1_1 = sy.factory1400.toPlainText()
    k = return_specific_terms(cursor, new_date, '2#智慧梁厂', '浇筑')
    m = return_specific_terms(cursor, new_date, '2#智慧梁厂', '安装')
    s_2 = return_specific_terms_qmx(cursor, new_date, '2#智慧梁厂', '湿接缝名称')
    f_2 = return_specific_terms_qmx(cursor, new_date, '2#智慧梁厂', '防撞护栏名称')
    # print('f_2')
    # print(f_2)
    q_2 = return_specific_terms_qmx(cursor, new_date, '2#智慧梁厂', '桥面铺装名称')
    hui_2 = k + m + s_2 + f_2 + q_2
    hui_2_2 = hui_2.replace("  ", "、").replace(" ", "")
    if hui_2_2 == '':
        hui_2_2 = '无'
    # note_information_2=input('\n2#梁厂今日完成详情：'+hui_2_2)
    sy.factory2400.setText(hui_2_2)
    hui_2_2 = sy.factory2400.toPlainText()
    note_information_all = '1#梁厂今日完成' + hui_1_1 + '\n' + '2#梁厂今日完成' + hui_2_2
    sht_2.range('f3').value = note_information_all

    sy.button2.setEnabled(True)
    sy.button1.setEnabled(False)


def main_2(cursor, sht_2, date, sy):
    global wb_1, excel_app
    sy.progressBar.setValue(10)
    sy.button1.setEnabled(False)
    print(date)
    new_date = date.replace('.', '/')
    sy.progressBar.setValue(20)
    # 填写梁片今日数量
    sht_2.range('c11').value = int(float(extract_data(new_date, new_date, '1#智慧梁厂', '25m', '梁片预制数量', cursor)) + float(
        extract_data(new_date, new_date, '12标', '25m', '梁片预制数量', cursor)))
    sht_2.range('c12').value = int(float(extract_data(new_date, new_date, '1#智慧梁厂', '40m', '梁片预制数量', cursor)) + float(
        extract_data(new_date, new_date, '12标', '40m', '梁片预制数量', cursor)))
    sht_2.range('e11').value = extract_data(new_date, new_date, '2#智慧梁厂', '25m', '梁片预制数量', cursor)
    sht_2.range('e12').value = extract_data(new_date, new_date, '2#智慧梁厂', '40m', '梁片预制数量', cursor)
    sht_2.range('c13').value = int(float(extract_data(new_date, new_date, '1#智慧梁厂', '25m', '梁片安装数量', cursor)) + float(
        extract_data(new_date, new_date, '12标', '25m', '梁片安装数量', cursor)))
    sht_2.range('c14').value = int(float(extract_data(new_date, new_date, '1#智慧梁厂', '40m', '梁片安装数量', cursor)) + float(
        extract_data(new_date, new_date, '12标', '40m', '梁片安装数量', cursor)))
    sht_2.range('e13').value = extract_data(new_date, new_date, '2#智慧梁厂', '25m', '梁片安装数量', cursor)
    sht_2.range('e14').value = extract_data(new_date, new_date, '2#智慧梁厂', '40m', '梁片安装数量', cursor)
    # 梁片产值
    sht_2.range('d11').value = float(extract_data(new_date, new_date, '1#智慧梁厂', '25m', '梁片预制产值', cursor)) + float(
        extract_data(new_date, new_date, '12标', '25m', '梁片预制产值', cursor))
    sht_2.range('d12').value = float(extract_data(new_date, new_date, '1#智慧梁厂', '40m', '梁片预制产值', cursor)) + float(
        extract_data(new_date, new_date, '12标', '40m', '梁片预制产值', cursor))
    sht_2.range('f11').value = extract_data(new_date, new_date, '2#智慧梁厂', '25m', '梁片预制产值', cursor)
    sht_2.range('f12').value = extract_data(new_date, new_date, '2#智慧梁厂', '40m', '梁片预制产值', cursor)
    sht_2.range('d13').value = float(extract_data(new_date, new_date, '1#智慧梁厂', '25m', '梁片安装产值', cursor)) + float(
        extract_data(new_date, new_date, '12标', '25m', '梁片安装产值', cursor))
    sht_2.range('d14').value = float(extract_data(new_date, new_date, '1#智慧梁厂', '40m', '梁片安装产值', cursor)) + float(
        extract_data(new_date, new_date, '12标', '40m', '梁片安装产值', cursor))
    sht_2.range('f13').value = extract_data(new_date, new_date, '2#智慧梁厂', '25m', '梁片安装产值', cursor)
    sht_2.range('f14').value = extract_data(new_date, new_date, '2#智慧梁厂', '40m', '梁片安装产值', cursor)
    # 桥面系数量
    sy.progressBar.setValue(30)
    sht_2.range('c15').value = str(float(extract_data(new_date, new_date, '1#智慧梁厂', '0', '湿接缝长度', cursor)) + float(
        extract_data(new_date, new_date, '12标', '0', '湿接缝长度', cursor)))
    sht_2.range('c16').value = str(float(extract_data(new_date, new_date, '1#智慧梁厂', '0', '防撞护栏长度', cursor)) + float(
        extract_data(new_date, new_date, '12标', '0', '防撞护栏长度', cursor)))
    sht_2.range('c17').value = str(float(extract_data(new_date, new_date, '1#智慧梁厂', '0', '桥面铺装面积', cursor)) + float(
        extract_data(new_date, new_date, '12标', '0', '桥面铺装面积', cursor)))
    sht_2.range('e15').value = extract_data(new_date, new_date, '2#智慧梁厂', '0', '湿接缝长度', cursor)
    sht_2.range('e16').value = extract_data(new_date, new_date, '2#智慧梁厂', '0', '防撞护栏长度', cursor)
    sht_2.range('e17').value = extract_data(new_date, new_date, '2#智慧梁厂', '0', '桥面铺装面积', cursor)
    # 桥面系产值
    sht_2.range('d15').value = str(float(extract_data(new_date, new_date, '1#智慧梁厂', '0', '湿接缝产值', cursor)) + float(
        extract_data(new_date, new_date, '12标', '0', '湿接缝产值', cursor)))
    sht_2.range('d16').value = str(float(extract_data(new_date, new_date, '1#智慧梁厂', '0', '防撞护栏产值', cursor)) + float(
        extract_data(new_date, new_date, '12标', '0', '防撞护栏产值', cursor)))
    sht_2.range('d17').value = str(float(extract_data(new_date, new_date, '1#智慧梁厂', '0', '桥面铺装产值', cursor)) + float(
        extract_data(new_date, new_date, '12标', '0', '桥面铺装产值', cursor)))
    sht_2.range('f15').value = extract_data(new_date, new_date, '2#智慧梁厂', '0', '湿接缝产值', cursor)
    sht_2.range('f16').value = extract_data(new_date, new_date, '2#智慧梁厂', '0', '防撞护栏产值', cursor)
    sht_2.range('f17').value = extract_data(new_date, new_date, '2#智慧梁厂', '0', '桥面铺装产值', cursor)
    sy.progressBar.setValue(30)
    # 读取位置信息并在指定位置填入数据，填入总数量
    order_1 = sht_2.range('h37:p37').value
    order_2 = sht_2.range('h39:p39').value
    # print(order_1)
    sht_2.range(order_1[0]).value = int(extract_data('2020/1/1', new_date, '1#智慧梁厂', '25m', '梁片预制数量', cursor)) + int(
        extract_data('2020/1/1', new_date, '12标', '25m', '梁片预制数量', cursor))
    sht_2.range(order_1[1]).value = int(extract_data('2020/1/1', new_date, '1#智慧梁厂', '40m', '梁片预制数量', cursor)) + int(
        extract_data('2020/1/1', new_date, '12标', '40m', '梁片预制数量', cursor))
    sht_2.range(order_1[2]).value = extract_data('2020/1/1', new_date, '2#智慧梁厂', '25m', '梁片预制数量', cursor)
    sht_2.range(order_1[3]).value = extract_data('2020/1/1', new_date, '2#智慧梁厂', '40m', '梁片预制数量', cursor)

    sht_2.range(order_1[4]).value = int(extract_data('2020/1/1', new_date, '1#智慧梁厂', '25m', '梁片安装数量', cursor)) + int(
        extract_data('2020/1/1', new_date, '12标', '25m', '梁片安装数量', cursor))
    sht_2.range(order_1[5]).value = int(extract_data('2020/1/1', new_date, '1#智慧梁厂', '40m', '梁片安装数量', cursor)) + int(
        extract_data('2020/1/1', new_date, '12标', '40m', '梁片安装数量', cursor))
    sht_2.range(order_1[6]).value = extract_data('2020/1/1', new_date, '2#智慧梁厂', '25m', '梁片安装数量', cursor)
    sht_2.range(order_1[7]).value = extract_data('2020/1/1', new_date, '2#智慧梁厂', '40m', '梁片安装数量', cursor)
    sy.progressBar.setValue(40)
    sht_2.range(order_2[0]).value = float(extract_data('2020/1/1', new_date, '1#智慧梁厂', '0', '湿接缝长度', cursor)) + float(
        extract_data('2020/1/1', new_date, '12标', '0', '湿接缝长度', cursor))
    sht_2.range(order_2[1]).value = float(extract_data('2020/1/1', new_date, '1#智慧梁厂', '0', '桥面铺装面积', cursor)) + float(
        extract_data('2020/1/1', new_date, '12标', '0', '桥面铺装面积', cursor))
    sht_2.range(order_2[2]).value = float(extract_data('2020/1/1', new_date, '1#智慧梁厂', '0', '防撞护栏长度', cursor)) + float(
        extract_data('2020/1/1', new_date, '12标', '0', '防撞护栏长度', cursor))

    sht_2.range(order_2[3]).value = extract_data('2020/1/1', new_date, '2#智慧梁厂', '0', '湿接缝长度', cursor)
    sht_2.range(order_2[4]).value = extract_data('2020/1/1', new_date, '2#智慧梁厂', '0', '桥面铺装面积', cursor)
    sht_2.range(order_2[5]).value = extract_data('2020/1/1', new_date, '2#智慧梁厂', '0', '防撞护栏长度', cursor)
    sy.state.setText('日报生成中......')
    sy.progressBar.setValue(50)
    # 更新总产值
    fill_total_value(sht_2)
    # 更新表中日期
    fill_today(date, sht_2)
    # 更新日期和图表
    update_chart(sht_2, date)
    # 截取图片并存储
    excel_catch_screen(sht_2, sht_2.range('p37').value, date)
    # print('图片已生成。')
    sy.state.setText('图片已生成。')
    sy.progressBar.setValue(80)
    # 返回差值
    erroe_value = str(error_output(cursor, sht_2))
    wb_1.save()
    wb_1.close()
    sy.state.setText('工作簿已保存并关闭。')
    sy.progressBar.setValue(90)
    # print('工作簿已保存并关闭。')
    excel_app.quit()
    sy.state.setText('WPS已退出。')
    # print('WPS已退出。')
    sy.state.setText('日报已更新,请查收。')
    sy.progressBar.setValue(100)
    # x=input('日报已更新,请查收。')
    QMessageBox.about(sy, '提示', '日报已更新,请查收。\n'+'产值误差：'+erroe_value)
    sy.button2.setEnabled(False)


def main_3(date, sy):
    '''更新浇筑安装台账'''
    try:
        wb_2 = openpyxl.load_workbook('浇筑安装台账.xlsx')
        # 获取位置信息
        list_location = read_cells(wb_2['位置信息']['b1:b17'])
        # 连接数据库,找出指定日期的梁片信息
        cursor = accdb(list_location[0].replace('"', ''))
        date = sy.date.date().toString('yyyy/M/d')
        cursor.execute("select * from 梁片信息简单查询 where 浇筑日期=#" + date + "#;")
        list_target_data = cursor.fetchall()

        # 获取最大行数，并将新的浇筑梁片填入报表
        max_row = wb_2['浇筑安装总台账'].max_row
        sht_target = wb_2['浇筑安装总台账']
        print(max_row)
        # 读取表格中所有的梁片编码
        list_order_all = read_cells(sht_target[str(list_location[-1]) + '1:' + list_location[-1] + str(max_row)])
        alignment_center = Alignment(horizontal='center', vertical='center')

        for i in range(len(list_target_data)):
            if int(list_target_data[i][-1]) in list_order_all:
                pass
            else:
                for j in range(len(list_target_data[i])):
                    if j != 6:
                        sht_target.cell(row=max_row + i + 1, column=ord(list_location[j + 1]) - 64,
                                        value=list_target_data[i][j])
                    elif j == 6:
                        x = list_target_data[i][j].strftime('%Y/%#m/%#d')
                        sht_target.cell(row=max_row + i + 1, column=ord(list_location[j + 1]) - 64,
                                        value=x).alignment = alignment_center

        # 更新安装日期、产值
        cursor.execute("select 梁片编码,安装日期,安装产值 from 梁片信息简单查询 where 安装日期=#" + date + "#;")
        list_ins = cursor.fetchall()
        cursor.close()
        list_order_ins = read_cells(sht_target[str(list_location[-1]) + '1:' + list_location[-1] + str(max_row)])
        list_number = []
        for i in range(len(list_ins)):
            '''返回目标安装编号行号，根据行号更新数据'''
            x = list_order_ins.index(list_ins[i][0]) + 1
            list_number.append(x)
            y = list_ins[i][1].strftime('%Y/%#m/%#d')
            sht_target.cell(row=x, column=ord(list_location[9]) - 64, value=y).alignment = alignment_center
            sht_target.cell(row=x, column=ord(list_location[10]) - 64, value=list_ins[i][2])
        wb_2.save('浇筑安装台账.xlsx')
        QMessageBox.about(sy, '提示', '浇筑安装台账已更新,请查收。')

    except FileNotFoundError:
        sy.state.setText('未找到“浇筑安装台账.xlsx”!')


if __name__ == '__main__':
    # 加载主程序窗体
    app = QApplication([])
    sy = Sy()
    sy.show()
    sy.state.setText('请点击打开日报表。')
    sy.actionDaily_report.triggered.connect(lambda: open_file(sy))
    sy.button1.clicked.connect(lambda: main(sy))
    sy.button2.clicked.connect(lambda: main_2(cursor, sht_2, date, sy))
    sy.button3.clicked.connect(lambda: main_3(date, sy))
    sys.exit(app.exec_())
