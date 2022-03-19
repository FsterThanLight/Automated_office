import openpyxl
import pypyodbc
import sys
from openpyxl.styles import Alignment
from tqdm import tqdm


def accdb(name):
    """建立与数据库的连接，返回游标"""
    # 取得当前文件目录
    mdb = 'Driver={Microsoft Access Driver (*.mdb)};DBQ=' + name
    try:
        conn = pypyodbc.win_connect_mdb(mdb)
        cursor = conn.cursor()
        print('成功连接数据库！\n')
    except pypyodbc.Error:
        print('无法识别该数据库！已退出程序。')
        cursor = ''
        sys.exit()
    except IndexError:
        print('无法识别该数据库！已退出程序。')
        cursor = ''
        sys.exit()
    # 建立连接
    return cursor


def find_bridges_under_construction(cursor):
    """查找正在建造中的桥，返回桥名"""
    list_1 = []
    cursor.execute('select 桥梁名称 from 正在建造中的桥 ')
    x = cursor.fetchall()
    for i in x:
        y = i[0]
        list_1.append(y)
    return list_1


def find_beam(cursor, bridegs_under_construction, judge):
    """查找目标梁片"""
    list_all = []
    if judge == '预制':
        for i in bridegs_under_construction:
            cursor.execute("select 桥梁名称,梁片,总产值 from 梁片信息汇总" \
                           " where 桥梁名称='" + i + "' and 浇筑日期 is null")
            for i in cursor.fetchall():
                list_all.append(i)
    elif judge == '安装':
        for i in bridegs_under_construction:
            cursor.execute("select 桥梁名称,梁片,梁片长度 from 梁片信息汇总 " \
                           "where 桥梁名称='" + i + "' and 安装日期 is null")
            for i in cursor.fetchall():
                list_all.append(i)
    return list_all


def set_column_width(sht):
    """设置表格列宽"""
    sht.column_dimensions['A'].width = 12
    sht.column_dimensions['B'].width = 20
    sht.column_dimensions['C'].width = 50
    sht.column_dimensions['D'].width = 25
    sht.column_dimensions['E'].width = 8
    sht.column_dimensions['F'].width = 8
    sht.column_dimensions['G'].width = 16
    sht.column_dimensions['H'].width = 12


def fill_in_excel_cast(beams):
    """新建工作簿将梁片信息写入表格中"""
    workbook = openpyxl.Workbook()
    sht = workbook['Sheet']
    sht.title = '浇筑剩余'
    alignment_center = Alignment(horizontal='center', vertical='center')
    sht.cell(row=1, column=3, value='桥梁名称')
    sht.cell(row=1, column=4, value='梁片名称')
    sht.cell(row=1, column=7, value='梁片产值')
    set_column_width(sht)
    for i in tqdm(range(len(beams))):
        sht.cell(row=i + 2, column=1, value='桥梁工程').alignment = alignment_center
        sht.cell(row=i + 2, column=2, value='预制').alignment = alignment_center
        sht.cell(row=i + 2, column=3, value=beams[i][0]).alignment = alignment_center
        sht.cell(row=i + 2, column=4, value=beams[i][1]).alignment = alignment_center
        sht.cell(row=i + 2, column=5, value='片').alignment = alignment_center
        sht.cell(row=i + 2, column=6, value=1).alignment = alignment_center
        sht.cell(row=i + 2, column=7, value=beams[i][2]).alignment = alignment_center
    workbook.save('在建中桥梁未完成梁片.xlsx')


def fill_in_excel_ins(beams):
    """加载工作表将安装剩余信息填入表格中"""
    workbook = openpyxl.load_workbook('在建中桥梁未完成梁片.xlsx')
    sht = workbook.create_sheet('安装剩余')
    alignment_center = Alignment(horizontal='center', vertical='center')
    sht.cell(row=1, column=3, value='桥梁名称')
    sht.cell(row=1, column=4, value='梁片名称')
    sht.cell(row=1, column=7, value='梁片长度')
    set_column_width(sht)
    for i in tqdm(range(len(beams))):
        sht.cell(row=i + 2, column=1, value='桥梁工程').alignment = alignment_center
        sht.cell(row=i + 2, column=2, value='预制').alignment = alignment_center
        sht.cell(row=i + 2, column=3, value=beams[i][0]).alignment = alignment_center
        sht.cell(row=i + 2, column=4, value=beams[i][1]).alignment = alignment_center
        sht.cell(row=i + 2, column=5, value='片').alignment = alignment_center
        sht.cell(row=i + 2, column=6, value=1).alignment = alignment_center
        sht.cell(row=i + 2, column=7, value=beams[i][2]).alignment = alignment_center
    workbook.save('在建中桥梁未完成梁片.xlsx')


if __name__ == '__main__':
    name = input('请输入数据库路径：').replace('"', "")
    # name = "E:\\OneDrive - 123\\日报、周报、月报、季报\\日报系统\\19标台账\ZCB1-19 T梁台账.accdb"
    cursor = accdb(name)
    # 正在建造中的桥桥名
    bridges_under_construction = find_bridges_under_construction(cursor)
    beam_cast = find_beam(cursor, bridges_under_construction, '预制')
    beam_ins = find_beam(cursor, bridges_under_construction, '安装')
    # 将梁片信息填入新建表格中
    fill_in_excel_cast(beam_cast)
    fill_in_excel_ins(beam_ins)
    print('数据生成中...')
    print('已完成数据提取。')
